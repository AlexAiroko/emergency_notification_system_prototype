from fastapi import UploadFile
from openpyxl import load_workbook

from app.exceptions.contact_import import EmptyImportFileError
from app.services.contact_import.parsers.base import BaseContactParser


class ExcelParser(BaseContactParser):
    def parse(self, file: UploadFile) -> list[dict]:
        file.file.seek(0)
        
        workbook = load_workbook(file.file)
        sheet = workbook.active
        
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            raise EmptyImportFileError()

        headers = list(rows[0])
        
        self.validate_headers(headers)
        
        contacts = []
        
        for values in rows[1:]:
            contacts.append(dict(zip(headers, values)))

        return contacts
